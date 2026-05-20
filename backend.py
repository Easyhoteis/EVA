from fastapi import FastAPI, Request, UploadFile, File
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
import uvicorn, os, hashlib, secrets, json, asyncio, csv, io, random, requests
from datetime import datetime, date
from zoneinfo import ZoneInfo
from dotenv import load_dotenv
import psycopg2
from psycopg2.extras import RealDictCursor

load_dotenv()

PORT = int(os.getenv("PORT", 3000))
DATABASE_URL = os.getenv("DATABASE_URL")

if not DATABASE_URL:
    print("ERRO: DATABASE_URL não configurada!")
    print("Adicione a variável DATABASE_URL no Railway")
    exit(1)

ZAPI_INST_AT = os.getenv("ZAPI_INSTANCE", "")
ZAPI_TOK_AT = os.getenv("ZAPI_TOKEN", "")
ZAPI_CLI_AT = os.getenv("ZAPI_CLIENT_TOKEN", "")
UPLOAD = "/tmp/uploads"
os.makedirs(UPLOAD, exist_ok=True)

app = FastAPI()
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])
app.mount("/uploads", StaticFiles(directory=UPLOAD), name="uploads")

def db():
    conn = psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
    # Configura timezone para Brasília
    with conn.cursor() as c:
        c.execute("SET TIME ZONE 'America/Sao_Paulo'")
    return conn

def hash_pass(s): return hashlib.sha256(s.encode()).hexdigest()
def token(): return secrets.token_urlsafe(32)

def init():
    conn = db()
    c = conn.cursor()
    c.execute("CREATE TABLE IF NOT EXISTS usuarios (id SERIAL PRIMARY KEY, email TEXT UNIQUE NOT NULL, senha_hash TEXT NOT NULL, nome TEXT NOT NULL, perfil TEXT NOT NULL, whatsapp TEXT, ativo INTEGER DEFAULT 1, criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP, ultimo_login TIMESTAMP)")
    
    # Adiciona coluna whatsapp se não existir
    try:
        c.execute("ALTER TABLE usuarios ADD COLUMN whatsapp TEXT")
        conn.commit()
    except:
        conn.rollback()
    
    c.execute("CREATE TABLE IF NOT EXISTS sessoes (id SERIAL PRIMARY KEY, token TEXT UNIQUE NOT NULL, usuario_id INTEGER NOT NULL, criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP)")
    c.execute("CREATE TABLE IF NOT EXISTS logs (id SERIAL PRIMARY KEY, usuario_id INTEGER, acao TEXT NOT NULL, detalhes TEXT, criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP)")
    c.execute("CREATE TABLE IF NOT EXISTS conversas (id SERIAL PRIMARY KEY, numero_cliente TEXT NOT NULL, nome_cliente TEXT, motivo TEXT, status TEXT DEFAULT 'aberto', fechado_por_id INTEGER, fechado_por_nome TEXT, criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP, fechado_em TIMESTAMP, observacoes TEXT)")
    
    # Adiciona coluna motivo se não existir
    try:
        c.execute("ALTER TABLE conversas ADD COLUMN motivo TEXT")
        conn.commit()
    except:
        conn.rollback()
    
    c.execute("CREATE TABLE IF NOT EXISTS mensagens (id SERIAL PRIMARY KEY, conversa_id INTEGER NOT NULL, remetente TEXT NOT NULL, conteudo TEXT NOT NULL, usuario_nome TEXT, criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP)")
    c.execute("CREATE TABLE IF NOT EXISTS contatos (id SERIAL PRIMARY KEY, nome TEXT NOT NULL, numero TEXT UNIQUE NOT NULL, email TEXT, tags TEXT, observacoes TEXT, conhecimento_ia TEXT, responsaveis TEXT, criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP)")
    
    # Adiciona coluna responsaveis se não existir
    try:
        c.execute("ALTER TABLE contatos ADD COLUMN responsaveis TEXT")
        conn.commit()
    except:
        conn.rollback()
    
    c.execute("CREATE TABLE IF NOT EXISTS templates (id SERIAL PRIMARY KEY, nome TEXT NOT NULL, conteudo TEXT NOT NULL, categoria TEXT, criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP)")
    c.execute("CREATE TABLE IF NOT EXISTS campanhas (id SERIAL PRIMARY KEY, nome TEXT NOT NULL, template_id INTEGER, mensagem TEXT NOT NULL, imagem_url TEXT, total_contatos INTEGER DEFAULT 0, enviadas INTEGER DEFAULT 0, falhadas INTEGER DEFAULT 0, status TEXT DEFAULT 'pendente', criado_por_id INTEGER, criado_por_nome TEXT, criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP, iniciado_em TIMESTAMP, finalizado_em TIMESTAMP)")
    c.execute("CREATE TABLE IF NOT EXISTS envios (id SERIAL PRIMARY KEY, campanha_id INTEGER, contato_id INTEGER, numero TEXT, nome TEXT, status TEXT DEFAULT 'pendente', resposta TEXT, enviado_em TIMESTAMP)")
    c.execute("CREATE TABLE IF NOT EXISTS config_antiban (id INTEGER PRIMARY KEY DEFAULT 1, limite_diario INTEGER DEFAULT 100, limite_por_hora INTEGER DEFAULT 30, delay_min INTEGER DEFAULT 3, delay_max INTEGER DEFAULT 7, pausa_a_cada INTEGER DEFAULT 30, pausa_segundos INTEGER DEFAULT 60, horario_inicio TEXT DEFAULT '08:00', horario_fim TEXT DEFAULT '20:00', ativo INTEGER DEFAULT 1)")
    c.execute("CREATE TABLE IF NOT EXISTS envios_diarios (data TEXT PRIMARY KEY, total INTEGER DEFAULT 0)")
    c.execute("CREATE TABLE IF NOT EXISTS config_zapi (tipo TEXT PRIMARY KEY, instance_id TEXT, token TEXT, client_token TEXT)")
    c.execute("CREATE TABLE IF NOT EXISTS config_sistema (chave TEXT PRIMARY KEY, valor TEXT)")
    c.execute("CREATE TABLE IF NOT EXISTS notificacoes (id SERIAL PRIMARY KEY, hotel_nome TEXT, hotel_numero TEXT, usuario_id INTEGER, usuario_nome TEXT, mensagem_original TEXT, enviado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP)")
    c.execute("CREATE TABLE IF NOT EXISTS listas_transmissao (id SERIAL PRIMARY KEY, nome TEXT NOT NULL, hoteis_ids TEXT NOT NULL, criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP)")
    conn.commit()
    
    c.execute("SELECT COUNT(*) as count FROM usuarios")
    if c.fetchone()['count'] == 0:
        c.execute("INSERT INTO usuarios (email, senha_hash, nome, perfil) VALUES (%s, %s, %s, %s)", ("admin@easy.com", hash_pass("admin123"), "Administrador", "admin"))
        conn.commit()
    
    c.execute("SELECT COUNT(*) as count FROM config_antiban")
    if c.fetchone()['count'] == 0:
        c.execute("INSERT INTO config_antiban VALUES (1, 100, 30, 3, 7, 30, 60, '08:00', '20:00', 1)")
        conn.commit()
    
    # Tabela config notificação grupo
    c.execute("""CREATE TABLE IF NOT EXISTS config_notificacao (
        id INTEGER PRIMARY KEY DEFAULT 1,
        grupo_id TEXT,
        grupo_nome TEXT,
        ativo INTEGER DEFAULT 0,
        enviar_individual INTEGER DEFAULT 1
    )""")
    
    c.execute("SELECT COUNT(*) as count FROM config_notificacao")
    if c.fetchone()['count'] == 0:
        c.execute("INSERT INTO config_notificacao VALUES (1, '120363425856356188-group', 'Grupo Easy Hotéis', 1, 1)")
        conn.commit()
    else:
        # Atualiza registro existente
        c.execute("UPDATE config_notificacao SET grupo_id = '120363425856356188-group', grupo_nome = 'Grupo Easy Hotéis', ativo = 1, enviar_individual = 1 WHERE id = 1")
        conn.commit()
    
    c.execute("SELECT COUNT(*) as count FROM config_zapi WHERE tipo = 'atendimento'")
    if c.fetchone()['count'] == 0 and ZAPI_INST_AT:
        c.execute("INSERT INTO config_zapi VALUES (%s, %s, %s, %s)", ("atendimento", ZAPI_INST_AT, ZAPI_TOK_AT, ZAPI_CLI_AT))
        conn.commit()
    
    c.execute("SELECT COUNT(*) as count FROM config_sistema WHERE chave = 'robo_ativo'")
    if c.fetchone()['count'] == 0:
        c.execute("INSERT INTO config_sistema VALUES ('robo_ativo', '1')")
        conn.commit()
    
    c.execute("SELECT COUNT(*) as count FROM templates")
    if c.fetchone()['count'] == 0:
        temps = [
            ("Regional - Easy 10 Anos", "Olá {nome}! 🏨\n\nConfira os destaques da semana em [REGIÃO]:\n\n🌟 [Hotel 1]\n🌟 [Hotel 2]\n🌟 [Hotel 3]\n\n🎉 Comemorando 10 anos da Easy, use o cupom EASY10 e ganhe 10% OFF!\n\nReserve já: [link]\n\n---\nNão quer mais receber? Responda SAIR", "regional"),
            ("Lançamento - Hotéis Parceiros", "Olá {nome}! ✨\n\nNOVOS PARCEIROS chegaram à Easy!\n\n🆕 [Hotel 1]\n🆕 [Hotel 2]\n🆕 [Hotel 3]\n\n🎊 10 anos Easy: cupom EASY10 = 10% OFF\n\nConheça: [link]\n\n---\nNão quer mais receber? Responda SAIR", "lancamento"),
            ("Promo - Ofertas Exclusivas", "Olá {nome}! 🔥\n\nOFERTAS IMPERDÍVEIS:\n\n💰 [Hotel 1] - de R$X por R$Y\n💰 [Hotel 2] - de R$X por R$Y\n\n✨ Use EASY10 e ganhe MAIS 10% OFF\n🎉 10 anos da Easy\n\nAproveite: [link]\n\n---\nNão quer mais receber? Responda SAIR", "promocao"),
            ("Aniversário 10 Anos", "Olá {nome}! 🎉\n\nA EASY ESTÁ DE PARABÉNS! 🎂\n\n10 anos transformando viagens em experiências!\n\n🎁 Cupom EASY10 com 10% OFF\nEm TODOS os hotéis parceiros!\n\nUse agora: [link]\n\nObrigado! ❤️\n\n---\nNão quer mais receber? Responda SAIR", "aniversario"),
        ]
        c.executemany("INSERT INTO templates (nome, conteudo, categoria) VALUES (%s, %s, %s)", temps)
        conn.commit()
    c.close()
    conn.close()

def valid_token(tok):
    if not tok: return None
    conn = db()
    c = conn.cursor()
    c.execute("SELECT u.id, u.email, u.nome, u.perfil, u.ativo FROM sessoes s JOIN usuarios u ON s.usuario_id = u.id WHERE s.token = %s", (tok,))
    r = c.fetchone()
    c.close()
    conn.close()
    return dict(r) if r and r['ativo'] else None

def get_user(req):
    auth = req.headers.get("Authorization", "")
    tok = auth.replace("Bearer ", "") if auth else None
    return valid_token(tok)

def log_acao(uid, acao, det=""):
    conn = db()
    c = conn.cursor()
    c.execute("INSERT INTO logs (usuario_id, acao, detalhes) VALUES (%s, %s, %s)", (uid, acao, det))
    conn.commit()
    c.close()
    conn.close()

def get_zapi(tipo="atendimento"):
    conn = db()
    c = conn.cursor()
    c.execute("SELECT * FROM config_zapi WHERE tipo = %s", (tipo,))
    r = c.fetchone()
    c.close()
    conn.close()
    return dict(r) if r else None

def enviar(num, msg, tipo="atendimento"):
    cfg = get_zapi(tipo)
    if not cfg: return {"ok": False}
    n = num.replace("+","").replace(" ","").replace("-","")
    url = f"https://api.z-api.io/instances/{cfg['instance_id']}/token/{cfg['token']}/send-text"
    try:
        r = requests.post(url, json={"phone": n, "message": msg}, headers={"Content-Type": "application/json", "Client-Token": cfg['client_token']}, timeout=30)
        return {"ok": r.status_code == 200}
    except: return {"ok": False}

def enviar_img(num, img, leg="", tipo="disparos"):
    cfg = get_zapi(tipo)
    if not cfg: return {"ok": False}
    n = num.replace("+","").replace(" ","").replace("-","")
    url = f"https://api.z-api.io/instances/{cfg['instance_id']}/token/{cfg['token']}/send-image"
    payload = {"phone": n, "image": img}
    if leg: payload["caption"] = leg
    try:
        r = requests.post(url, json=payload, headers={"Content-Type": "application/json", "Client-Token": cfg['client_token']}, timeout=60)
        return {"ok": r.status_code == 200}
    except: return {"ok": False}

def status_zapi(tipo="atendimento"):
    cfg = get_zapi(tipo)
    if not cfg: return {"conectado": False}
    try:
        r = requests.get(f"https://api.z-api.io/instances/{cfg['instance_id']}/token/{cfg['token']}/status", headers={"Client-Token": cfg['client_token']}, timeout=10)
        return {"conectado": r.status_code == 200, "data": r.json() if r.status_code == 200 else None}
    except: return {"conectado": False}

def cfg_antiban():
    conn = db()
    c = conn.cursor()
    c.execute("SELECT * FROM config_antiban WHERE id = 1")
    r = c.fetchone()
    c.close()
    conn.close()
    return dict(r) if r else None

def envios_hj():
    conn = db()
    c = conn.cursor()
    c.execute("SELECT total FROM envios_diarios WHERE data = %s", (date.today().isoformat(),))
    r = c.fetchone()
    c.close()
    conn.close()
    return r['total'] if r else 0

def add_envio():
    conn = db()
    c = conn.cursor()
    hj = date.today().isoformat()
    c.execute("INSERT INTO envios_diarios (data, total) VALUES (%s, 1) ON CONFLICT(data) DO UPDATE SET total = envios_diarios.total + 1", (hj,))
    conn.commit()
    c.close()
    conn.close()

def pode_enviar():
    cfg = cfg_antiban()
    if not cfg or not cfg['ativo']: return True, "ok"
    agora = datetime.now().strftime("%H:%M")
    if agora < cfg['horario_inicio'] or agora > cfg['horario_fim']: return False, "horario"
    if envios_hj() >= cfg['limite_diario']: return False, "limite"
    return True, "ok"

def now_br():
    """Retorna datetime atual no horário de Brasília"""
    return datetime.now(ZoneInfo("America/Sao_Paulo"))

def robo_on():
    conn = db()
    c = conn.cursor()
    c.execute("SELECT valor FROM config_sistema WHERE chave = 'robo_ativo'")
    r = c.fetchone()
    c.close()
    conn.close()
    return r['valor'] == '1' if r else True

def toggle_robo(on):
    conn = db()
    c = conn.cursor()
    v = '1' if on else '0'
    c.execute("INSERT INTO config_sistema (chave, valor) VALUES ('robo_ativo', %s) ON CONFLICT(chave) DO UPDATE SET valor = %s", (v, v))
    conn.commit()
    c.close()
    conn.close()
    return True

def ia(msg, conhec="", hotel="Hotel"):
    # Menu fixo - SEM IA - 100% confiável
    msg_lower = msg.lower().strip()
    
    # Cliente escolheu opção 1, 2 ou 3
    if msg_lower == '1':
        return """✅ Fecho de disponibilidade registrado!

Envie os detalhes em UMA mensagem:
📅 Data(s)
🛏️ Categoria(s)

Exemplo: 14/05 suite master"""
    
    elif msg_lower == '2':
        return """✅ Atualização de tarifas registrada!

Envie os detalhes em UMA mensagem:
📅 Período
🛏️ Categoria(s)
💰 Novos valores

Exemplo: 10/05 a 15/05 suite executiva R$350"""
    
    elif msg_lower == '3':
        return """✅ Solicitação registrada!

Descreva o que precisa que um atendente vai ajudar em breve."""
    
    # Se mensagem tem palavras-chave de detalhes - confirma recebimento
    palavras_detalhes = ['suite', 'standard', 'luxo', 'executiv', 'casal', 'solteiro', 'master', '/', 'r$', 'tarif', 'duplo', 'triplo']
    if any(palavra in msg_lower for palavra in palavras_detalhes):
        return "✅ Detalhes recebidos! Atendente vai processar em breve."
    
    # Primeira mensagem ou não reconheceu - mostra menu
    return f"""Olá! Sou a EVA, assistente da Easy Hotéis! 😊

Como posso ajudar?

1️⃣ Fechar disponibilidade
2️⃣ Atualizar tarifas
3️⃣ Outros assuntos

Digite o número da opção."""

@app.get("/")
async def root(): return {"ok": True}

@app.get("/painel")
async def painel(): return FileResponse("painel.html")

@app.post("/api/login")
async def login(req: Request):
    d = await req.json()
    email = d.get("email", "").lower().strip()
    senha = d.get("senha", "")
    if not email or not senha: return JSONResponse({"sucesso": False, "erro": "Campos vazios"}, 400)
    conn = db()
    c = conn.cursor()
    c.execute("SELECT * FROM usuarios WHERE email = %s AND ativo = 1", (email,))
    u = c.fetchone()
    if not u or u['senha_hash'] != hash_pass(senha): 
        c.close()
        conn.close()
        return JSONResponse({"sucesso": False, "erro": "Email ou senha incorretos"}, 401)
    tok = token()
    c.execute("INSERT INTO sessoes (token, usuario_id) VALUES (%s, %s)", (tok, u['id']))
    c.execute("UPDATE usuarios SET ultimo_login = CURRENT_TIMESTAMP WHERE id = %s", (u['id'],))
    conn.commit()
    c.close()
    conn.close()
    log_acao(u['id'], "login", "ok")
    return {"sucesso": True, "token": tok, "usuario": {"id": u['id'], "nome": u['nome'], "email": u['email'], "perfil": u['perfil']}}

@app.post("/api/logout")
async def logout(req: Request):
    tok = req.headers.get("Authorization", "").replace("Bearer ", "")
    if tok:
        conn = db()
        c = conn.cursor()
        c.execute("DELETE FROM sessoes WHERE token = %s", (tok,))
        conn.commit()
        c.close()
        conn.close()
    return {"sucesso": True}

@app.get("/api/me")
async def me(req: Request):
    u = get_user(req)
    return {"sucesso": True, "usuario": u} if u else JSONResponse({"sucesso": False}, 401)

@app.get("/api/usuarios")
async def list_users(req: Request):
    u = get_user(req)
    if not u or u['perfil'] != 'admin': return JSONResponse({"sucesso": False}, 403)
    conn = db()
    c = conn.cursor()
    c.execute("SELECT id, email, nome, perfil, ativo, criado_em, ultimo_login FROM usuarios ORDER BY nome")
    rows = c.fetchall()
    c.close()
    conn.close()
    return {"sucesso": True, "usuarios": [dict(r) for r in rows]}

@app.post("/api/usuarios")
async def criar_user(req: Request):
    u = get_user(req)
    if not u or u['perfil'] != 'admin': return JSONResponse({"sucesso": False}, 403)
    d = await req.json()
    email = d.get("email", "").lower().strip()
    senha = d.get("senha", "")
    nome = d.get("nome", "")
    perfil = d.get("perfil", "atendente")
    whatsapp = d.get("whatsapp", "")
    if not email or not senha or not nome: return JSONResponse({"sucesso": False, "erro": "Campos vazios"}, 400)
    if perfil not in ["admin", "atendente", "marketing"]: return JSONResponse({"sucesso": False, "erro": "Perfil invalido"}, 400)
    conn = db()
    c = conn.cursor()
    try:
        c.execute("INSERT INTO usuarios (email, senha_hash, nome, perfil, whatsapp) VALUES (%s, %s, %s, %s, %s)", (email, hash_pass(senha), nome, perfil, whatsapp))
        conn.commit()
        log_acao(u['id'], "criar_user", f"{email}")
        c.close()
        conn.close()
        return {"sucesso": True}
    except Exception as e:
        c.close()
        conn.close()
        print(f"ERRO AO CRIAR USUARIO: {str(e)}")
        return JSONResponse({"sucesso": False, "erro": str(e)}, 400)

@app.put("/api/usuarios/{uid}")
async def update_user(uid: int, req: Request):
    u = get_user(req)
    if not u or u['perfil'] != 'admin': return JSONResponse({"sucesso": False}, 403)
    d = await req.json()
    print(f"UPDATE USUARIO {uid}: payload recebido = {d}")
    conn = db()
    c = conn.cursor()
    campos, valores = [], []
    if "nome" in d: campos.append("nome = %s"); valores.append(d["nome"])
    if "email" in d: campos.append("email = %s"); valores.append(d["email"].lower().strip())
    if "perfil" in d and d["perfil"] in ["admin", "atendente", "marketing"]: campos.append("perfil = %s"); valores.append(d["perfil"])
    if "ativo" in d: campos.append("ativo = %s"); valores.append(1 if d["ativo"] else 0)
    if "senha" in d and d["senha"]: campos.append("senha_hash = %s"); valores.append(hash_pass(d["senha"]))
    if "whatsapp" in d: 
        campos.append("whatsapp = %s")
        valores.append(d["whatsapp"])
        print(f"WhatsApp a ser salvo: {d['whatsapp']}")
    if campos:
        valores.append(uid)
        sql = f"UPDATE usuarios SET {', '.join(campos)} WHERE id = %s"
        print(f"SQL: {sql}")
        print(f"Valores: {valores}")
        c.execute(sql, valores)
        conn.commit()
        print(f"UPDATE executado! Linhas afetadas: {c.rowcount}")
    c.close()
    conn.close()
    return {"sucesso": True}

@app.delete("/api/usuarios/{uid}")
async def del_user(uid: int, req: Request):
    u = get_user(req)
    if not u or u['perfil'] != 'admin': return JSONResponse({"sucesso": False}, 403)
    if uid == u['id']: return JSONResponse({"sucesso": False, "erro": "Nao pode deletar a si mesmo"}, 400)
    conn = db()
    c = conn.cursor()
    c.execute("DELETE FROM usuarios WHERE id = %s", (uid,))
    c.execute("DELETE FROM sessoes WHERE usuario_id = %s", (uid,))
    conn.commit()
    c.close()
    conn.close()
    return {"sucesso": True}

@app.get("/api/config/zapi")
async def get_cfg_zapi(req: Request):
    u = get_user(req)
    if not u or u['perfil'] != 'admin': return JSONResponse({"sucesso": False}, 403)
    conn = db()
    c = conn.cursor()
    c.execute("SELECT * FROM config_zapi")
    rows = c.fetchall()
    c.close()
    conn.close()
    return {"sucesso": True, "configs": {r['tipo']: dict(r) for r in rows}}

@app.post("/api/config/zapi")
async def save_cfg_zapi(req: Request):
    u = get_user(req)
    if not u or u['perfil'] != 'admin': return JSONResponse({"sucesso": False}, 403)
    d = await req.json()
    tipo = d.get("tipo")
    if tipo not in ["atendimento", "disparos"]: return JSONResponse({"sucesso": False}, 400)
    conn = db()
    c = conn.cursor()
    c.execute("INSERT INTO config_zapi (tipo, instance_id, token, client_token) VALUES (%s, %s, %s, %s) ON CONFLICT(tipo) DO UPDATE SET instance_id = %s, token = %s, client_token = %s",
        (tipo, d.get("instance_id"), d.get("token"), d.get("client_token"), d.get("instance_id"), d.get("token"), d.get("client_token")))
    conn.commit()
    c.close()
    conn.close()
    return {"sucesso": True}

@app.post("/webhook/zapi")
async def webhook(req: Request):
    d = await req.json()
    if d.get("fromMe"): return {"success": True}
    num = d.get("phone")
    nome = d.get("senderName", "")
    msg = d.get("text", {}).get("message") if d.get("text") else None
    if not num or not msg: return {"success": False}
    
    # verifica se é resposta de atendente (numero cadastrado como whatsapp de usuario)
    conn = db()
    c = conn.cursor()
    c.execute("SELECT id, nome FROM usuarios WHERE whatsapp = %s", (num,))
    atend = c.fetchone()
    
    if atend:
        # é atendente respondendo! processa como conclusão
        msg_lower = msg.lower()
        palavras_conclusao = ['concluido', 'concluído', 'feito', 'realizado', 'ok', 'pronto', 'done', '✅', 'resolvido']
        
        if any(p in msg_lower for p in palavras_conclusao):
            # busca ultima notificacao desse atendente pra saber qual hotel
            c.execute("SELECT hotel_nome, hotel_numero, mensagem_original FROM notificacoes WHERE usuario_id = %s ORDER BY enviado_em DESC LIMIT 1", (atend['id'],))
            notif = c.fetchone()
            
            if notif:
                hotel_num = notif['hotel_numero']
                hotel_nome = notif['hotel_nome']
                
                # responde pro hotel
                msg_conclusao = f"""✅ *Concluído!*

Sua solicitação foi processada com sucesso por nossa equipe.

Qualquer dúvida, estamos à disposição! 😊"""
                enviar(hotel_num, msg_conclusao, "atendimento")
                
                # fecha a conversa
                c.execute("SELECT id FROM conversas WHERE numero_cliente = %s AND status = 'aberto' LIMIT 1", (hotel_num,))
                conv = c.fetchone()
                if conv:
                    c.execute("UPDATE conversas SET status = 'fechado', fechado_em = CURRENT_TIMESTAMP, fechado_por_id = %s, fechado_por_nome = %s WHERE id = %s",
                        (atend['id'], atend['nome'], conv['id']))
                    c.execute("INSERT INTO mensagens (conversa_id, remetente, conteudo) VALUES (%s, 'eva', %s)", (conv['id'], msg_conclusao))
                    conn.commit()
                
                # responde pro atendente confirmando
                enviar(num, f"✅ Confirmado! Avisei o {hotel_nome} que foi concluído e fechei o atendimento.", "atendimento")
            
            c.close()
            conn.close()
            return {"success": True, "tipo": "conclusao_atendente"}
    
    # verifica se conversa foi fechada (cliente agradecendo depois)
    c.execute("SELECT id FROM conversas WHERE numero_cliente = %s AND status = 'fechado' ORDER BY fechado_em DESC LIMIT 1", (num,))
    conv_fechada = c.fetchone()
    
    if conv_fechada:
        msg_lower = msg.lower()
        palavras_agradecimento = ['obrigado', 'obrigada', 'valeu', 'thanks', 'vlw', 'brigadão', '👍', '🙏']
        
        if any(p in msg_lower for p in palavras_agradecimento):
            # É agradecimento - responde e fecha de novo
            c.execute("UPDATE conversas SET status = 'aberto' WHERE id = %s", (conv_fechada['id'],))
            c.execute("INSERT INTO mensagens (conversa_id, remetente, conteudo) VALUES (%s, 'cliente', %s)", (conv_fechada['id'], msg))
            conn.commit()
            
            resp = "De nada! Estamos sempre à disposição. 😊"
            c.execute("INSERT INTO mensagens (conversa_id, remetente, conteudo) VALUES (%s, 'eva', %s)", (conv_fechada['id'], resp))
            c.execute("UPDATE conversas SET status = 'fechado', fechado_em = CURRENT_TIMESTAMP WHERE id = %s", (conv_fechada['id'],))
            conn.commit()
            c.close()
            conn.close()
            enviar(num, resp, "atendimento")
            return {"success": True, "tipo": "agradecimento_pos_fechamento"}
        else:
            # Não é agradecimento - abre conversa nova (não reabre a antiga)
            # Deixa cair no código normal abaixo que vai criar conversa nova
            pass
    
    if not robo_on():
        c.execute("SELECT id FROM conversas WHERE numero_cliente = %s AND status = 'aberto' LIMIT 1", (num,))
        r = c.fetchone()
        cid = r['id'] if r else None
        if not cid:
            c.execute("INSERT INTO conversas (numero_cliente, nome_cliente) VALUES (%s, %s) RETURNING id", (num, nome))
            cid = c.fetchone()['id']
        c.execute("INSERT INTO mensagens (conversa_id, remetente, conteudo) VALUES (%s, %s, %s)", (cid, "cliente", msg))
        conn.commit()
        c.close()
        conn.close()
        return {"success": True, "robo": "pausado"}
    
    c.execute("SELECT id FROM conversas WHERE numero_cliente = %s AND status = 'aberto' LIMIT 1", (num,))
    r = c.fetchone()
    cid = r['id'] if r else None
    sessao_nova = False
    if not cid:
        c.execute("INSERT INTO conversas (numero_cliente, nome_cliente) VALUES (%s, %s) RETURNING id", (num, nome))
        cid = c.fetchone()['id']
        sessao_nova = True
    
    c.execute("SELECT conhecimento_ia, nome, responsaveis FROM contatos WHERE numero = %s", (num,))
    cont = c.fetchone()
    conhec = cont['conhecimento_ia'] if cont and cont['conhecimento_ia'] else ""
    hotel = cont['nome'] if cont else nome
    responsaveis_json = cont['responsaveis'] if cont else None
    
    # Detecta motivo (primeiro por número, depois por palavra-chave)
    motivo = None
    msg_lower = msg.lower().strip()
    
    if msg_lower == '1':
        motivo = 'Fecho de Disponibilidade'
    elif msg_lower == '2':
        motivo = 'Atualização de Tarifas'
    elif msg_lower == '3':
        motivo = 'Outros Assuntos'
    elif 'fech' in msg_lower or 'disponibilidade' in msg_lower:
        motivo = 'Fecho de Disponibilidade'
    elif 'tarif' in msg_lower or 'preço' in msg_lower or 'valor' in msg_lower:
        motivo = 'Atualização de Tarifas'
    elif 'bloqu' in msg_lower:
        motivo = 'Bloqueio de Quartos'
    
    c.execute("INSERT INTO mensagens (conversa_id, remetente, conteudo) VALUES (%s, %s, %s)", (cid, "cliente", msg))
    
    if motivo:
        c.execute("UPDATE conversas SET motivo = %s WHERE id = %s", (motivo, cid))
    
    conn.commit()
    
    # Conta mensagens ANTES de responder pra decidir se responde
    c.execute("SELECT criado_em FROM conversas WHERE id = %s", (cid,))
    conv_criado = c.fetchone()['criado_em']
    
    c.execute("SELECT COUNT(*) as count FROM mensagens WHERE conversa_id = %s AND remetente = 'cliente' AND criado_em >= %s", (cid, conv_criado))
    num_msgs = c.fetchone()['count']
    
    # SÓ RESPONDE SE ESTÁ EM CONTATOS E É 1ª, 2ª OU 3ª MENSAGEM
    if cont and num_msgs <= 3:
        resp = ia(msg, conhec, hotel)
        c.execute("INSERT INTO mensagens (conversa_id, remetente, conteudo) VALUES (%s, %s, %s)", (cid, "eva", resp))
        conn.commit()
        enviar(num, resp, "atendimento")
    
    # Conta mensagens do cliente desde que a conversa foi CRIADA/REABERTA (atualiza contagem)
    c.execute("SELECT criado_em FROM conversas WHERE id = %s", (cid,))
    conv_criado = c.fetchone()['criado_em']
    
    # Conta mensagens do cliente DESDE a criação desta sessão
    c.execute("SELECT COUNT(*) as count FROM mensagens WHERE conversa_id = %s AND remetente = 'cliente' AND criado_em >= %s", (cid, conv_criado))
    num_msgs = c.fetchone()['count']
    print(f"DEBUG: Número de mensagens do cliente (sessão atual): {num_msgs}")
    print(f"DEBUG: Sessão criada em: {conv_criado}")
    
    # SÓ ENVIA NOTIFICAÇÃO NA TERCEIRA MENSAGEM DA SESSÃO ATUAL
    if num_msgs == 3:
        print("DEBUG: Terceira mensagem! Enviando notificação...")
        # Pega TODAS as mensagens do cliente
        c.execute("SELECT conteudo FROM mensagens WHERE conversa_id = %s AND remetente = 'cliente' ORDER BY id", (cid,))
        todas_msgs = c.fetchall()
        
        # Formata mensagens (substitui número por texto)
        msgs_formatadas = []
        for m in todas_msgs:
            conteudo = m['conteudo'].strip()
            if conteudo == '1':
                msgs_formatadas.append('"1 - Fecho de Disponibilidade"')
            elif conteudo == '2':
                msgs_formatadas.append('"2 - Atualização de Tarifas"')
            elif conteudo == '3':
                msgs_formatadas.append('"3 - Outros Assuntos"')
            else:
                msgs_formatadas.append(f'"{conteudo}"')
        
        msgs_texto = '\n'.join(msgs_formatadas)
        
        # Pega motivo atualizado
        c.execute("SELECT motivo FROM conversas WHERE id = %s", (cid,))
        conv = c.fetchone()
        motivo_atual = conv['motivo'] if conv and conv['motivo'] else 'N/A'
        
        # Verifica config de notificação em grupo
        c.execute("SELECT * FROM config_notificacao WHERE id = 1")
        config_notif = c.fetchone()
        print(f"DEBUG: Config notificação: ativo={config_notif['ativo'] if config_notif else 'None'}, grupo_id={config_notif['grupo_id'] if config_notif else 'None'}")
        
        # Envia notificação no grupo (se configurado)
        if config_notif and config_notif['ativo'] and config_notif['grupo_id']:
            print(f"ENVIANDO NOTIFICAÇÃO NO GRUPO: {config_notif['grupo_nome']}")
            msg_grupo = f"""🔔 *NOVO PEDIDO*

🏨 *Hotel:* {hotel}
📱 *Número:* {num}
⏰ *Horário:* {now_br().strftime('%H:%M')}

📝 *Mensagem:*
{msgs_texto}

🎯 *Motivo:* {motivo_atual}

👉 *Acesse:* https://eva-easyhoteis-83036260b078.herokuapp.com/painel"""
            
            try:
                enviar(config_notif['grupo_id'], msg_grupo, "atendimento")
                print(f"✅ Notificação enviada no grupo")
            except Exception as e:
                print(f"ERRO ENVIAR GRUPO: {str(e)}")
        
        # Envia notificação individual (se configurado)
        if responsaveis_json and (not config_notif or config_notif.get('enviar_individual', 1)):
            print(f"NOTIFICANDO RESPONSÁVEIS: {responsaveis_json}")
            try:
                resp_ids = json.loads(responsaveis_json)
                print(f"IDs responsáveis: {resp_ids}")
                if resp_ids:
                    c.execute("SELECT id, nome, whatsapp FROM usuarios WHERE id = ANY(%s) AND whatsapp IS NOT NULL", (resp_ids,))
                    usuarios = c.fetchall()
                    print(f"Usuários encontrados: {len(usuarios)}")
                    for u in usuarios:
                        print(f"Enviando notificação para {u['nome']} ({u['whatsapp']})")
                        msg_notif = f"""🔔 *NOVO PEDIDO - {hotel}*

Cliente: {hotel}
Número: {num}

Mensagem:
{msgs_texto}

Responda aqui quando concluir!"""
                        enviar(u['whatsapp'], msg_notif, "atendimento")
                        c.execute("INSERT INTO notificacoes (hotel_nome, hotel_numero, usuario_id, usuario_nome, mensagem_original) VALUES (%s, %s, %s, %s, %s)",
                            (hotel, num, u['id'], u['nome'], msg))
                        conn.commit()
            except Exception as e:
                print(f"ERRO NOTIFICAÇÃO: {str(e)}")
        else:
            print("SEM RESPONSÁVEIS PARA NOTIFICAR")
    
    c.close()
    conn.close()
    return {"success": True}

@app.get("/api/conversas/abertas")
async def conv_abertas(req: Request):
    u = get_user(req)
    if not u: return JSONResponse({"sucesso": False}, 401)
    conn = db()
    c = conn.cursor()
    c.execute("SELECT c.*, (SELECT COUNT(*) FROM mensagens WHERE conversa_id = c.id) as total_mensagens, (SELECT conteudo FROM mensagens WHERE conversa_id = c.id ORDER BY criado_em DESC LIMIT 1) as ultima_mensagem FROM conversas c WHERE status = 'aberto' ORDER BY criado_em DESC")
    rows = c.fetchall()
    
    convs = []
    for r in rows:
        conv = dict(r)
        conv['minha_responsabilidade'] = False
        c.execute("SELECT responsaveis FROM contatos WHERE numero = %s", (r['numero_cliente'],))
        cont = c.fetchone()
        if cont and cont['responsaveis']:
            try:
                resp_ids = json.loads(cont['responsaveis'])
                if u['id'] in resp_ids:
                    conv['minha_responsabilidade'] = True
            except: pass
        convs.append(conv)
    
    c.close()
    conn.close()
    return {"sucesso": True, "conversas": convs, "total": len(convs)}

@app.get("/api/conversas/fechadas")
async def conv_fechadas(req: Request):
    u = get_user(req)
    if not u: return JSONResponse({"sucesso": False}, 401)
    conn = db()
    c = conn.cursor()
    c.execute("SELECT c.*, (SELECT COUNT(*) FROM mensagens WHERE conversa_id = c.id) as total_mensagens FROM conversas c WHERE status = 'fechado' ORDER BY fechado_em DESC")
    rows = c.fetchall()
    c.close()
    conn.close()
    return {"sucesso": True, "conversas": [dict(r) for r in rows], "total": len(rows)}

@app.get("/api/conversas/{cid}/mensagens")
async def msgs(cid: int, req: Request):
    u = get_user(req)
    if not u: return JSONResponse({"sucesso": False}, 401)
    conn = db()
    c = conn.cursor()
    c.execute("SELECT * FROM mensagens WHERE conversa_id = %s ORDER BY criado_em", (cid,))
    rows = c.fetchall()
    c.close()
    conn.close()
    return {"sucesso": True, "mensagens": [dict(r) for r in rows]}

@app.post("/api/conversas/{cid}/fechar")
async def fechar(cid: int, req: Request):
    u = get_user(req)
    if not u: return JSONResponse({"sucesso": False}, 401)
    d = await req.json()
    conn = db()
    c = conn.cursor()
    c.execute("SELECT numero_cliente FROM conversas WHERE id = %s", (cid,))
    r = c.fetchone()
    num = r['numero_cliente'] if r else None
    c.execute("UPDATE conversas SET status = 'fechado', fechado_em = CURRENT_TIMESTAMP, fechado_por_id = %s, fechado_por_nome = %s, observacoes = %s WHERE id = %s",
        (u['id'], u['nome'], d.get("observacoes", ""), cid))
    conn.commit()
    
    # Só envia mensagem de fechamento se está em contatos
    if num:
        c.execute("SELECT id FROM contatos WHERE numero = %s", (num,))
        if c.fetchone():
            enviar(num, f"*{u['nome']}:*\nObrigado! Atendimento encerrado. 😊", "atendimento")
    
    c.close()
    conn.close()
    return {"sucesso": True}

@app.post("/api/conversas/{cid}/responder")
async def responder(cid: int, req: Request):
    u = get_user(req)
    if not u: return JSONResponse({"sucesso": False}, 401)
    d = await req.json()
    msg = d.get("mensagem", "")
    if not msg: return JSONResponse({"sucesso": False}, 400)
    conn = db()
    c = conn.cursor()
    c.execute("SELECT numero_cliente FROM conversas WHERE id = %s", (cid,))
    r = c.fetchone()
    if not r:
        c.close()
        conn.close()
        return JSONResponse({"sucesso": False}, 404)
    num = r['numero_cliente']
    msg_final = f"*{u['nome']}:*\n{msg}"
    c.execute("INSERT INTO mensagens (conversa_id, remetente, conteudo, usuario_nome) VALUES (%s, %s, %s, %s)", (cid, "atendente", msg_final, u['nome']))
    conn.commit()
    c.close()
    conn.close()
    return {"sucesso": True, "envio": enviar(num, msg_final, "atendimento")}

@app.get("/api/contatos")
async def list_contatos(req: Request):
    u = get_user(req)
    if not u: return JSONResponse({"sucesso": False}, 401)
    conn = db()
    c = conn.cursor()
    c.execute("SELECT * FROM contatos ORDER BY nome")
    rows = c.fetchall()
    c.close()
    conn.close()
    return {"sucesso": True, "contatos": [dict(r) for r in rows], "total": len(rows)}

@app.post("/api/contatos")
async def criar_contato(req: Request):
    u = get_user(req)
    if not u or u['perfil'] not in ['admin', 'marketing']: return JSONResponse({"sucesso": False}, 403)
    d = await req.json()
    if not d.get("nome") or not d.get("numero"): return JSONResponse({"sucesso": False}, 400)
    conn = db()
    c = conn.cursor()
    try:
        resp_ids = d.get("responsaveis", [])
        resp_json = json.dumps(resp_ids) if resp_ids else None
        c.execute("INSERT INTO contatos (nome, numero, email, tags, observacoes, conhecimento_ia, responsaveis) VALUES (%s, %s, %s, %s, %s, %s, %s)",
            (d["nome"], d["numero"], d.get("email", ""), d.get("tags", ""), d.get("observacoes", ""), d.get("conhecimento_ia", ""), resp_json))
        conn.commit()
        c.close()
        conn.close()
        return {"sucesso": True}
    except Exception as e:
        c.close()
        conn.close()
        erro_msg = str(e)
        if "duplicate key" in erro_msg or "unique constraint" in erro_msg:
            return JSONResponse({"sucesso": False, "erro": "Numero ja cadastrado"}, 400)
        else:
            print(f"ERRO CRIAR CONTATO: {erro_msg}")
            return JSONResponse({"sucesso": False, "erro": f"Erro ao salvar: {erro_msg}"}, 400)

@app.put("/api/contatos/{cid}")
async def editar_contato(cid: int, req: Request):
    u = get_user(req)
    if not u or u['perfil'] not in ['admin', 'marketing']: return JSONResponse({"sucesso": False}, 403)
    d = await req.json()
    if not d.get("nome") or not d.get("numero"): return JSONResponse({"sucesso": False}, 400)
    conn = db()
    c = conn.cursor()
    try:
        resp_ids = d.get("responsaveis", [])
        resp_json = json.dumps(resp_ids) if resp_ids else None
        c.execute("UPDATE contatos SET nome = %s, numero = %s, email = %s, tags = %s, conhecimento_ia = %s, responsaveis = %s WHERE id = %s",
            (d["nome"], d["numero"], d.get("email", ""), d.get("tags", ""), d.get("conhecimento_ia", ""), resp_json, cid))
        conn.commit()
        c.close()
        conn.close()
        return {"sucesso": True}
    except Exception as e:
        c.close()
        conn.close()
        print(f"ERRO EDITAR CONTATO: {str(e)}")
        return JSONResponse({"sucesso": False, "erro": str(e)}, 400)

@app.delete("/api/contatos/{cid}")
async def del_contato(cid: int, req: Request):
    u = get_user(req)
    if not u or u['perfil'] not in ['admin', 'marketing']: return JSONResponse({"sucesso": False}, 403)
    conn = db()
    c = conn.cursor()
    c.execute("DELETE FROM contatos WHERE id = %s", (cid,))
    conn.commit()
    c.close()
    conn.close()
    return {"sucesso": True}

@app.post("/api/contatos/importar")
async def importar(file: UploadFile = File(...)):
    content = await file.read()
    text = content.decode('utf-8')
    reader = csv.DictReader(io.StringIO(text))
    conn = db()
    c = conn.cursor()
    imp, err = 0, 0
    for row in reader:
        nome = row.get('nome') or row.get('Nome') or row.get('NOME')
        numero = row.get('numero') or row.get('Numero') or row.get('NUMERO') or row.get('telefone')
        email = row.get('email') or row.get('Email') or ""
        if nome and numero:
            try:
                c.execute("INSERT INTO contatos (nome, numero, email) VALUES (%s, %s, %s) ON CONFLICT DO NOTHING", (nome, numero, email))
                if c.rowcount > 0: imp += 1
                else: err += 1
            except: err += 1
    conn.commit()
    c.close()
    conn.close()
    return {"sucesso": True, "importados": imp, "erros": err}

@app.post("/api/contatos/preview-csv")
async def preview(file: UploadFile = File(...)):
    content = await file.read()
    text = content.decode('utf-8')
    reader = csv.DictReader(io.StringIO(text))
    conts = []
    for row in reader:
        nome = row.get('nome') or row.get('Nome') or row.get('NOME')
        numero = row.get('numero') or row.get('Numero') or row.get('NUMERO') or row.get('telefone')
        if nome and numero: conts.append({"nome": nome, "numero": numero})
    return {"sucesso": True, "contatos": conts, "total": len(conts)}

@app.post("/api/upload/imagem")
async def upload(file: UploadFile = File(...)):
    ext = file.filename.split(".")[-1].lower()
    if ext not in ["jpg", "jpeg", "png", "webp"]: return JSONResponse({"sucesso": False}, 400)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    nome = f"camp_{ts}.{ext}"
    path = os.path.join(UPLOAD, nome)
    content = await file.read()
    with open(path, "wb") as f: f.write(content)
    return {"sucesso": True, "url": f"/uploads/{nome}"}

@app.get("/api/templates")
async def list_temps(req: Request):
    u = get_user(req)
    if not u: return JSONResponse({"sucesso": False}, 401)
    conn = db()
    c = conn.cursor()
    c.execute("SELECT * FROM templates ORDER BY nome")
    rows = c.fetchall()
    c.close()
    conn.close()
    return {"sucesso": True, "templates": [dict(r) for r in rows]}

@app.post("/api/templates")
async def criar_temp(req: Request):
    u = get_user(req)
    if not u or u['perfil'] not in ['admin', 'marketing']: return JSONResponse({"sucesso": False}, 403)
    d = await req.json()
    if not d.get("nome") or not d.get("conteudo"): return JSONResponse({"sucesso": False}, 400)
    conn = db()
    c = conn.cursor()
    c.execute("INSERT INTO templates (nome, conteudo, categoria) VALUES (%s, %s, %s)", (d["nome"], d["conteudo"], d.get("categoria", "geral")))
    conn.commit()
    c.close()
    conn.close()
    return {"sucesso": True}

@app.delete("/api/templates/{tid}")
async def del_temp(tid: int, req: Request):
    u = get_user(req)
    if not u or u['perfil'] not in ['admin', 'marketing']: return JSONResponse({"sucesso": False}, 403)
    conn = db()
    c = conn.cursor()
    c.execute("DELETE FROM templates WHERE id = %s", (tid,))
    conn.commit()
    c.close()
    conn.close()
    return {"sucesso": True}

@app.get("/api/campanhas")
async def list_camps(req: Request):
    u = get_user(req)
    if not u: return JSONResponse({"sucesso": False}, 401)
    conn = db()
    c = conn.cursor()
    c.execute("SELECT * FROM campanhas ORDER BY criado_em DESC")
    rows = c.fetchall()
    c.close()
    conn.close()
    return {"sucesso": True, "campanhas": [dict(r) for r in rows]}

@app.post("/api/campanhas")
async def criar_camp(req: Request):
    u = get_user(req)
    if not u or u['perfil'] not in ['admin', 'marketing']: return JSONResponse({"sucesso": False}, 403)
    d = await req.json()
    nome = d.get("nome")
    msg = d.get("mensagem")
    cids = d.get("contato_ids", [])
    csvs = d.get("contatos_csv", [])
    img = d.get("imagem_url")
    if not nome or not msg: return JSONResponse({"sucesso": False}, 400)
    total = len(cids) + len(csvs)
    if total == 0: return JSONResponse({"sucesso": False, "erro": "Selecione contatos"}, 400)
    cfg = cfg_antiban()
    hj = envios_hj()
    if cfg and cfg['ativo']:
        disp = cfg['limite_diario'] - hj
        if total > disp: return JSONResponse({"sucesso": False, "erro": f"Limite {disp}"}, 400)
    conn = db()
    c = conn.cursor()
    c.execute("INSERT INTO campanhas (nome, template_id, mensagem, imagem_url, total_contatos, status, criado_por_id, criado_por_nome) VALUES (%s, %s, %s, %s, %s, 'pendente', %s, %s) RETURNING id",
        (nome, d.get("template_id"), msg, img, total, u['id'], u['nome']))
    cid = c.fetchone()['id']
    for cid_cont in cids:
        c.execute("SELECT nome, numero FROM contatos WHERE id = %s", (cid_cont,))
        r = c.fetchone()
        if r: c.execute("INSERT INTO envios (campanha_id, contato_id, numero, nome, status) VALUES (%s, %s, %s, %s, 'pendente')", (cid, cid_cont, r['numero'], r['nome']))
    for cont in csvs:
        c.execute("INSERT INTO envios (campanha_id, contato_id, numero, nome, status) VALUES (%s, NULL, %s, %s, 'pendente')", (cid, cont['numero'], cont['nome']))
    conn.commit()
    c.close()
    conn.close()
    asyncio.create_task(processar(cid))
    return {"sucesso": True, "campanha_id": cid}

async def processar(cid):
    conn = db()
    c = conn.cursor()
    c.execute("UPDATE campanhas SET status = 'enviando', iniciado_em = CURRENT_TIMESTAMP WHERE id = %s", (cid,))
    conn.commit()
    c.execute("SELECT mensagem, imagem_url FROM campanhas WHERE id = %s", (cid,))
    camp = c.fetchone()
    if not camp: 
        c.close()
        conn.close()
        return
    msg_base, img = camp['mensagem'], camp['imagem_url']
    cfg = cfg_antiban()
    img_url = None
    if img:
        base = os.getenv("RAILWAY_PUBLIC_DOMAIN", "web-production-69fb05.up.railway.app")
        img_url = f"https://{base}{img}" if base else None
    c.execute("SELECT id, numero, nome FROM envios WHERE campanha_id = %s AND status = 'pendente'", (cid,))
    envs = c.fetchall()
    env, falh = 0, 0
    for idx, e in enumerate(envs):
        eid, num, nm = e['id'], e['numero'], e['nome']
        ok, _ = pode_enviar()
        if not ok:
            c.execute("UPDATE envios SET status = 'pausado' WHERE id = %s", (eid,))
            conn.commit()
            continue
        if cfg and idx > 0 and idx % cfg['pausa_a_cada'] == 0: await asyncio.sleep(cfg['pausa_segundos'])
        msg_final = msg_base.replace("{nome}", nm or "").replace("{numero}", num or "")
        if img_url: result = enviar_img(num, img_url, msg_final, "disparos")
        else: result = enviar(num, msg_final, "disparos")
        if result.get("ok"):
            c.execute("UPDATE envios SET status = 'enviado', enviado_em = CURRENT_TIMESTAMP WHERE id = %s", (eid,))
            env += 1
            add_envio()
        else:
            c.execute("UPDATE envios SET status = 'falhou' WHERE id = %s", (eid,))
            falh += 1
        c.execute("UPDATE campanhas SET enviadas = %s, falhadas = %s WHERE id = %s", (env, falh, cid))
        conn.commit()
        delay = random.uniform(cfg['delay_min'], cfg['delay_max']) if cfg else 4
        await asyncio.sleep(delay)
    c.execute("UPDATE campanhas SET status = 'finalizada', finalizado_em = CURRENT_TIMESTAMP WHERE id = %s", (cid,))
    conn.commit()
    c.close()
    conn.close()

@app.get("/api/antiban/config")
async def get_antiban(req: Request):
    u = get_user(req)
    if not u: return JSONResponse({"sucesso": False}, 401)
    cfg = cfg_antiban()
    hj = envios_hj()
    return {"sucesso": True, "config": cfg, "envios_hoje": hj, "disponivel_hoje": (cfg['limite_diario'] - hj) if cfg else 0}

@app.post("/api/antiban/config")
async def save_antiban(req: Request):
    u = get_user(req)
    if not u or u['perfil'] != 'admin': return JSONResponse({"sucesso": False}, 403)
    d = await req.json()
    conn = db()
    c = conn.cursor()
    c.execute("UPDATE config_antiban SET limite_diario=%s, limite_por_hora=%s, delay_min=%s, delay_max=%s, pausa_a_cada=%s, pausa_segundos=%s, horario_inicio=%s, horario_fim=%s, ativo=%s WHERE id=1",
        (d.get("limite_diario",100), d.get("limite_por_hora",30), d.get("delay_min",3), d.get("delay_max",7), d.get("pausa_a_cada",30), d.get("pausa_segundos",60),
         d.get("horario_inicio","08:00"), d.get("horario_fim","20:00"), 1 if d.get("ativo",True) else 0))
    conn.commit()
    c.close()
    conn.close()
    return {"sucesso": True}

@app.get("/api/relatorios")
async def relat(req: Request):
    u = get_user(req)
    if not u: return JSONResponse({"sucesso": False}, 401)
    conn = db()
    c = conn.cursor()
    c.execute("SELECT COUNT(*) as count FROM conversas")
    total_conv = c.fetchone()['count']
    c.execute("SELECT COUNT(*) as count FROM conversas WHERE status='aberto'")
    abertos = c.fetchone()['count']
    c.execute("SELECT COUNT(*) as count FROM conversas WHERE status='fechado'")
    fechados = c.fetchone()['count']
    c.execute("SELECT COUNT(*) as count FROM contatos")
    conts = c.fetchone()['count']
    c.execute("SELECT COUNT(*) as count FROM campanhas")
    camps = c.fetchone()['count']
    c.execute("SELECT SUM(enviadas) as total FROM campanhas")
    env = c.fetchone()['total'] or 0
    c.execute("SELECT AVG(EXTRACT(EPOCH FROM (fechado_em - criado_em))/60) as tempo FROM conversas WHERE status='fechado'")
    tempo = c.fetchone()['tempo']
    tempo_med = int(tempo) if tempo else 0
    c.execute("SELECT fechado_por_nome, COUNT(*) as total FROM conversas WHERE status='fechado' AND fechado_por_nome IS NOT NULL GROUP BY fechado_por_nome ORDER BY total DESC LIMIT 5")
    top = [dict(r) for r in c.fetchall()]
    
    # Novos relatórios
    c.execute("SELECT nome_cliente, COUNT(*) as total FROM conversas WHERE nome_cliente IS NOT NULL GROUP BY nome_cliente ORDER BY total DESC LIMIT 10")
    top_hoteis = [dict(r) for r in c.fetchall()]
    
    c.execute("SELECT motivo, COUNT(*) as total FROM conversas WHERE motivo IS NOT NULL GROUP BY motivo ORDER BY total DESC")
    por_motivo = [dict(r) for r in c.fetchall()]
    
    c.close()
    conn.close()
    return {"sucesso": True, "relatorios": {"totalConversas": total_conv, "abertos": abertos, "fechados": fechados, "tempoMedioMinutos": tempo_med,
        "totalContatos": conts, "totalCampanhas": camps, "totalEnviadas": env, "enviadosHoje": envios_hj(), "topAtendentes": top,
        "topHoteis": top_hoteis, "porMotivo": por_motivo}}

@app.get("/api/relatorios/export")
async def export_excel(req: Request):
    u = get_user(req)
    if not u or u['perfil'] != 'admin': return JSONResponse({"sucesso": False}, 403)
    
    import io
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    conn = db()
    c = conn.cursor()
    
    c.execute("""
        SELECT c.id, c.nome_cliente, c.numero_cliente, c.motivo, c.status, 
               c.fechado_por_nome, c.criado_em, c.fechado_em,
               EXTRACT(EPOCH FROM (c.fechado_em - c.criado_em))/60 as tempo_min
        FROM conversas c
        ORDER BY c.criado_em DESC
    """)
    conversas = c.fetchall()
    
    c.close()
    conn.close()
    
    # Cria workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório EVA"
    
    # Cores EVA (roxo/azul)
    cor_header = "6C63FF"  # Roxo EVA
    cor_header_text = "FFFFFF"  # Branco
    cor_linha_par = "F5F5FF"  # Azul clarinho
    
    # Headers
    headers = ["ID", "Hotel", "Número", "Motivo", "Status", "Atendente", "Abertura", "Fechamento", "Tempo (min)"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color=cor_header_text, size=12)
        cell.fill = PatternFill(start_color=cor_header, end_color=cor_header, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Dados
    for idx, conv in enumerate(conversas, 2):
        tempo = int(conv['tempo_min']) if conv['tempo_min'] else 0
        abertura = conv['criado_em'].strftime('%d/%m/%Y %H:%M') if conv['criado_em'] else ''
        fechamento = conv['fechado_em'].strftime('%d/%m/%Y %H:%M') if conv['fechado_em'] else ''
        
        row_data = [
            conv['id'],
            conv['nome_cliente'] or conv['numero_cliente'],
            conv['numero_cliente'],
            conv['motivo'] or 'N/A',
            conv['status'],
            conv['fechado_por_nome'] or 'N/A',
            abertura,
            fechamento,
            tempo
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(idx, col, value)
            cell.alignment = Alignment(horizontal="center" if col in [1, 5, 9] else "left", vertical="center")
            cell.border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            # Linhas alternadas
            if idx % 2 == 0:
                cell.fill = PatternFill(start_color=cor_linha_par, end_color=cor_linha_par, fill_type="solid")
    
    # Ajusta largura das colunas
    ws.column_dimensions['A'].width = 8   # ID
    ws.column_dimensions['B'].width = 25  # Hotel
    ws.column_dimensions['C'].width = 18  # Número
    ws.column_dimensions['D'].width = 25  # Motivo
    ws.column_dimensions['E'].width = 12  # Status
    ws.column_dimensions['F'].width = 20  # Atendente
    ws.column_dimensions['G'].width = 18  # Abertura
    ws.column_dimensions['H'].width = 18  # Fechamento
    ws.column_dimensions['I'].width = 14  # Tempo
    
    # Congela primeira linha
    ws.freeze_panes = "A2"
    
    # Salva em memória
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    import base64
    excel_data = base64.b64encode(output.read()).decode()
    
    return JSONResponse({
        "sucesso": True,
        "filename": f"relatorio_eva_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        "data": excel_data
    })

@app.get("/api/zapi/status")
async def zapi_st(req: Request):
    return {"atendimento": status_zapi("atendimento"), "disparos": status_zapi("disparos")}

@app.get("/api/zapi/grupos")
async def listar_grupos(req: Request):
    u = get_user(req)
    if not u or u['perfil'] != 'admin': return JSONResponse({"sucesso": False}, 403)
    
    # Tenta endpoint list-chats
    url = f"https://api.z-api.io/instances/{ZAPI_INST_AT}/token/{ZAPI_TOK_AT}/list-chats"
    
    try:
        print(f"Buscando grupos na Z-API: {url}")
        res = requests.get(url, timeout=15)
        print(f"Status Z-API: {res.status_code}")
        
        if res.status_code != 200:
            # Se list-chats não funcionar, tenta contacts
            url2 = f"https://api.z-api.io/instances/{ZAPI_INST_AT}/token/{ZAPI_TOK_AT}/contacts"
            print(f"Tentando endpoint alternativo: {url2}")
            res = requests.get(url2, timeout=15)
            print(f"Status Z-API (contacts): {res.status_code}")
            
            if res.status_code != 200:
                return JSONResponse({"sucesso": False, "erro": f"Z-API retornou status {res.status_code}. Verifique se a instância está conectada."}, 500)
        
        data = res.json()
        print(f"Resposta Z-API: {type(data)}")
        
        grupos = []
        
        # Processa resposta
        if isinstance(data, dict):
            chats = data.get('chats', []) or data.get('contacts', [])
        else:
            chats = data
        
        print(f"Total de chats: {len(chats)}")
        
        for chat in chats:
            try:
                # ID do chat
                chat_id = chat.get('id', '')
                
                # Verifica se é grupo (termina com @g.us)
                if '@g.us' in str(chat_id):
                    nome = chat.get('name', '') or chat.get('pushname', '') or chat.get('formattedTitle', '') or 'Sem nome'
                    
                    # Tenta pegar número de participantes
                    num_participantes = 0
                    if 'groupMetadata' in chat:
                        num_participantes = len(chat['groupMetadata'].get('participants', []))
                    
                    grupos.append({
                        'id': chat_id,
                        'nome': nome,
                        'participantes': num_participantes
                    })
                    print(f"Grupo encontrado: {nome} ({chat_id})")
                    
            except Exception as e2:
                print(f"Erro ao processar chat: {str(e2)}")
                continue
        
        print(f"Total de grupos encontrados: {len(grupos)}")
        
        if len(grupos) == 0:
            return JSONResponse({"sucesso": False, "erro": "Nenhum grupo encontrado. Verifique se o número está em algum grupo no WhatsApp."}, 400)
        
        return {"sucesso": True, "grupos": grupos}
        
    except Exception as e:
        print(f"ERRO LISTAR GRUPOS: {str(e)}")
        import traceback
        traceback.print_exc()
        return JSONResponse({"sucesso": False, "erro": f"Erro ao conectar com Z-API: {str(e)}"}, 500)

@app.get("/api/notificacao/config")
async def get_config_notificacao(req: Request):
    u = get_user(req)
    if not u or u['perfil'] != 'admin': return JSONResponse({"sucesso": False}, 403)
    conn = db()
    c = conn.cursor()
    c.execute("SELECT * FROM config_notificacao WHERE id = 1")
    config = c.fetchone()
    c.close()
    conn.close()
    return {"sucesso": True, "config": dict(config) if config else {}}

@app.post("/api/notificacao/config")
async def salvar_config_notificacao(req: Request):
    u = get_user(req)
    if not u or u['perfil'] != 'admin': return JSONResponse({"sucesso": False}, 403)
    d = await req.json()
    conn = db()
    c = conn.cursor()
    c.execute("UPDATE config_notificacao SET grupo_id = %s, grupo_nome = %s, ativo = %s, enviar_individual = %s WHERE id = 1",
        (d.get('grupo_id'), d.get('grupo_nome'), 1 if d.get('ativo') else 0, 1 if d.get('enviar_individual') else 0))
    conn.commit()
    c.close()
    conn.close()
    return {"sucesso": True}

@app.get("/api/listas")
async def get_listas(req: Request):
    u = get_user(req)
    if not u or u['perfil'] != 'admin': return JSONResponse({"sucesso": False}, 403)
    conn = db()
    c = conn.cursor()
    c.execute("SELECT * FROM listas_transmissao ORDER BY nome")
    listas = [dict(r) for r in c.fetchall()]
    c.close()
    conn.close()
    return {"sucesso": True, "listas": listas}

@app.post("/api/listas")
async def criar_lista(req: Request):
    u = get_user(req)
    if not u or u['perfil'] != 'admin': return JSONResponse({"sucesso": False}, 403)
    d = await req.json()
    nome = d.get("nome", "")
    hoteis_ids = d.get("hoteis_ids", [])
    if not nome or not hoteis_ids:
        return JSONResponse({"sucesso": False, "erro": "Nome e hotéis obrigatórios"}, 400)
    conn = db()
    c = conn.cursor()
    c.execute("INSERT INTO listas_transmissao (nome, hoteis_ids) VALUES (%s, %s)", (nome, json.dumps(hoteis_ids)))
    conn.commit()
    c.close()
    conn.close()
    return {"sucesso": True}

@app.put("/api/listas/{lid}")
async def atualizar_lista(lid: int, req: Request):
    u = get_user(req)
    if not u or u['perfil'] != 'admin': return JSONResponse({"sucesso": False}, 403)
    d = await req.json()
    nome = d.get("nome", "")
    hoteis_ids = d.get("hoteis_ids", [])
    if not nome or not hoteis_ids:
        return JSONResponse({"sucesso": False, "erro": "Nome e hotéis obrigatórios"}, 400)
    conn = db()
    c = conn.cursor()
    c.execute("UPDATE listas_transmissao SET nome = %s, hoteis_ids = %s WHERE id = %s", (nome, json.dumps(hoteis_ids), lid))
    conn.commit()
    c.close()
    conn.close()
    return {"sucesso": True}

@app.delete("/api/listas/{lid}")
async def deletar_lista(lid: int, req: Request):
    u = get_user(req)
    if not u or u['perfil'] != 'admin': return JSONResponse({"sucesso": False}, 403)
    conn = db()
    c = conn.cursor()
    c.execute("DELETE FROM listas_transmissao WHERE id = %s", (lid,))
    conn.commit()
    c.close()
    conn.close()
    return {"sucesso": True}

@app.post("/api/transmissao")
async def lista_transmissao(req: Request):
    u = get_user(req)
    if not u or u['perfil'] != 'admin': return JSONResponse({"sucesso": False, "erro": "Apenas Admin pode usar Lista de Transmissão"}, 403)
    
    d = await req.json()
    mensagem = d.get("mensagem", "")
    lista_id = d.get("lista_id", None)
    
    if not mensagem:
        return JSONResponse({"sucesso": False, "erro": "Mensagem vazia"}, 400)
    
    if not lista_id:
        return JSONResponse({"sucesso": False, "erro": "Selecione uma lista"}, 400)
    
    conn = db()
    c = conn.cursor()
    
    # Busca lista
    c.execute("SELECT hoteis_ids FROM listas_transmissao WHERE id = %s", (lista_id,))
    lista = c.fetchone()
    if not lista:
        c.close()
        conn.close()
        return JSONResponse({"sucesso": False, "erro": "Lista não encontrada"}, 400)
    
    hoteis_ids = json.loads(lista['hoteis_ids'])
    
    # Busca hotéis
    placeholders = ','.join(['%s'] * len(hoteis_ids))
    c.execute(f"SELECT nome, numero FROM contatos WHERE id IN ({placeholders}) ORDER BY nome", tuple(hoteis_ids))
    hoteis = c.fetchall()
    c.close()
    conn.close()
    
    if not hoteis:
        return JSONResponse({"sucesso": False, "erro": "Nenhum hotel na lista"}, 400)
    
    # Envia para todos
    enviados = 0
    falhados = 0
    
    for hotel in hoteis:
        msg_final = f"📢 *AVISO EASY HOTÉIS*\n\n{mensagem}\n\n---\nMensagem enviada para lista selecionada."
        result = enviar(hotel['numero'], msg_final, "atendimento")
        
        if result.get("ok"):
            enviados += 1
            await asyncio.sleep(2)  # Delay entre envios
        else:
            falhados += 1
    
    return {"sucesso": True, "enviados": enviados, "falhados": falhados, "total": len(hoteis)}

@app.get("/api/sistema/robo")
async def get_robo(req: Request):
    u = get_user(req)
    if not u: return JSONResponse({"sucesso": False}, 401)
    return {"sucesso": True, "ativo": robo_on()}

@app.post("/api/sistema/robo")
async def toggle_robo_api(req: Request):
    u = get_user(req)
    if not u or u['perfil'] != 'admin': return JSONResponse({"sucesso": False}, 403)
    d = await req.json()
    on = d.get("ativo", True)
    ok = toggle_robo(on)
    if ok: log_acao(u['id'], "toggle_robo", f"{'on' if on else 'off'}")
    return {"sucesso": ok, "ativo": on}

@app.get("/api/notificacoes")
async def list_notifs(req: Request):
    u = get_user(req)
    if not u: return JSONResponse({"sucesso": False}, 401)
    conn = db()
    c = conn.cursor()
    if u['perfil'] == 'admin':
        c.execute("SELECT * FROM notificacoes ORDER BY enviado_em DESC LIMIT 100")
    else:
        c.execute("SELECT * FROM notificacoes WHERE usuario_id = %s ORDER BY enviado_em DESC LIMIT 100", (u['id'],))
    rows = c.fetchall()
    c.close()
    conn.close()
    return {"sucesso": True, "notificacoes": [dict(r) for r in rows]}

if __name__ == "__main__":
    print("="*40)
    print("EVA v5 - PostgreSQL")
    print("="*40)
    init()
    print(f"Porta: {PORT}")
    print("Login: admin@easy.com / admin123")
    print("="*40)
    uvicorn.run(app, host="0.0.0.0", port=PORT)
